"""Excel export for the property financial report."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font

from apps.reservations.reports.exports._formatting import (
    EXCEL_DATE_NUMBER_FORMAT,
    EXCEL_DATETIME_NUMBER_FORMAT,
    EXCEL_SHEET_TITLE,
    display_booking_reference,
    display_external_reference,
    export_filename,
    format_guest_names,
    format_period_iso,
    write_excel_money_cell,
)
from apps.reservations.reports.types import PropertyFinancialReportResult

HEADER_ROW = 8
HEADERS = (
    "Booking code",
    "External ID",
    "Check-in",
    "Check-out",
    "Rooms",
    "Nights",
    "Gross",
    "Commission",
    "Net",
    "Currency",
    "Source",
    "Guests",
    "Payout status",
    "Payout date",
)


def _write_excel_date_cell(cell, value) -> None:
    cell.value = value
    cell.number_format = EXCEL_DATE_NUMBER_FORMAT


def render_property_financial_report_xlsx(result: PropertyFinancialReportResult) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXCEL_SHEET_TITLE

    meta = result.meta
    bold = Font(bold=True)

    sheet["A1"] = "Property financial report"
    sheet["A1"].font = bold
    sheet["A2"] = "Property"
    sheet["B2"] = meta.property_name
    sheet["A3"] = "Period (check-out)"
    sheet["B3"] = format_period_iso(meta.check_out_from, meta.check_out_to)
    sheet["A4"] = "Generated at"
    generated_cell = sheet["B4"]
    generated_cell.value = meta.generated_at.replace(tzinfo=None)
    generated_cell.number_format = EXCEL_DATETIME_NUMBER_FORMAT
    sheet["A5"] = "Currency"
    sheet["B5"] = meta.currency
    if meta.rows_with_missing_commission:
        sheet["A6"] = "Rows with missing commission"
        sheet["B6"] = meta.rows_with_missing_commission
    if meta.rows_without_confirmed_payout:
        sheet["A7"] = "Rows without confirmed payout"
        sheet["B7"] = meta.rows_without_confirmed_payout

    for column, title in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=column, value=title)
        cell.font = bold

    for offset, row in enumerate(result.rows, start=1):
        excel_row = HEADER_ROW + offset
        sheet.cell(row=excel_row, column=1, value=display_booking_reference(
            booking_code=row.booking_code,
            external_id=row.external_id,
        ))
        sheet.cell(
            row=excel_row,
            column=2,
            value=display_external_reference(
                booking_code=row.booking_code,
                external_id=row.external_id,
            ),
        )
        _write_excel_date_cell(sheet.cell(row=excel_row, column=3), row.check_in)
        _write_excel_date_cell(sheet.cell(row=excel_row, column=4), row.check_out)
        sheet.cell(row=excel_row, column=5, value=", ".join(row.room_labels))
        sheet.cell(row=excel_row, column=6, value=row.nights)
        write_excel_money_cell(
            sheet.cell(row=excel_row, column=7),
            row.gross,
            null_as_zero=True,
        )
        write_excel_money_cell(sheet.cell(row=excel_row, column=8), row.commission)
        write_excel_money_cell(sheet.cell(row=excel_row, column=9), row.net)
        sheet.cell(row=excel_row, column=10, value=row.currency)
        sheet.cell(row=excel_row, column=11, value=row.source)
        sheet.cell(row=excel_row, column=12, value=format_guest_names(row.guests))
        sheet.cell(row=excel_row, column=13, value=row.payout_status.value)
        payout_date_cell = sheet.cell(row=excel_row, column=14)
        if row.payout_received_at is not None:
            _write_excel_date_cell(payout_date_cell, row.payout_received_at)
        else:
            payout_date_cell.value = None

    totals_row = HEADER_ROW + len(result.rows) + 2
    totals = result.totals
    sheet.cell(row=totals_row, column=1, value="Totals").font = bold
    sheet.cell(row=totals_row, column=6, value=totals.nights)
    write_excel_money_cell(sheet.cell(row=totals_row, column=7), totals.gross)
    write_excel_money_cell(sheet.cell(row=totals_row, column=8), totals.commission)
    write_excel_money_cell(sheet.cell(row=totals_row, column=9), totals.net)
    sheet.cell(row=totals_row + 1, column=1, value="Reservation count")
    sheet.cell(row=totals_row + 1, column=2, value=totals.reservation_count)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def property_financial_report_xlsx_filename(result: PropertyFinancialReportResult) -> str:
    return export_filename(result, "xlsx")
