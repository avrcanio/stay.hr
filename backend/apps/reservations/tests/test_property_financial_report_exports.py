import io
import numbers
import zipfile
from datetime import date, datetime

from django.test import SimpleTestCase
from openpyxl import load_workbook

from apps.reservations.reports.exports._formatting import (
    EXCEL_MONEY_NUMBER_FORMAT,
    EXCEL_SHEET_TITLE,
    export_filename,
)
from apps.reservations.reports.exports.excel import HEADERS, HEADER_ROW, render_property_financial_report_xlsx
from apps.reservations.reports.exports.pdf import (
    property_financial_report_pdf_filename,
    render_property_financial_report_html,
    render_property_financial_report_pdf,
)
from apps.reservations.tests.fixtures.property_financial_report_result import (
    sample_property_financial_report_result,
)


class PropertyFinancialReportExportTests(SimpleTestCase):
    def setUp(self):
        self.result = sample_property_financial_report_result()

    def test_export_filename(self):
        self.assertEqual(
            export_filename(self.result, "pdf"),
            "property-financial-uzorita-2026-03-01_2026-03-31.pdf",
        )
        self.assertEqual(
            export_filename(self.result, "xlsx"),
            "property-financial-uzorita-2026-03-01_2026-03-31.xlsx",
        )

    def test_pdf_returns_non_empty_bytes(self):
        payload = render_property_financial_report_pdf(self.result)
        self.assertIsInstance(payload, bytes)
        self.assertGreater(len(payload), 100)
        self.assertTrue(payload.startswith(b"%PDF"))

    def test_pdf_html_includes_key_labels_and_hr_decimals(self):
        html = render_property_financial_report_html(self.result)
        self.assertIn("Financijski izvještaj", html)
        self.assertIn("BK-COMPLETE", html)
        self.assertIn("Ana Anić", html)
        self.assertIn("Uzorita Luxury Rooms", html)
        self.assertIn("150,00", html)
        self.assertIn("10.03.2026", html)

    def test_pdf_filename(self):
        self.assertEqual(
            property_financial_report_pdf_filename(self.result),
            "property-financial-uzorita-2026-03-01_2026-03-31.pdf",
        )

    def test_xlsx_returns_valid_workbook(self):
        payload = render_property_financial_report_xlsx(self.result)
        self.assertIsInstance(payload, bytes)
        self.assertGreater(len(payload), 100)
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            self.assertIn("xl/workbook.xml", archive.namelist())

    def test_xlsx_structure_headers_rows_and_totals(self):
        workbook = load_workbook(io.BytesIO(render_property_financial_report_xlsx(self.result)))
        sheet = workbook.active

        self.assertEqual(sheet.title, EXCEL_SHEET_TITLE)
        self.assertEqual(
            [cell.value for cell in sheet[HEADER_ROW]],
            list(HEADERS),
        )
        self.assertEqual(sheet.cell(row=HEADER_ROW + 1, column=1).value, "BK-COMPLETE")
        self.assertEqual(sheet.cell(row=HEADER_ROW + 2, column=1).value, "BK-NO-COMM")
        self.assertEqual(sheet.max_row, HEADER_ROW + len(self.result.rows) + 3)

        totals_row = HEADER_ROW + len(self.result.rows) + 2
        self.assertEqual(sheet.cell(row=totals_row, column=1).value, "Totals")
        self.assertEqual(sheet.cell(row=totals_row, column=6).value, self.result.totals.nights)
        self.assertEqual(sheet.cell(row=totals_row, column=7).value, self.result.totals.gross)
        self.assertEqual(sheet.cell(row=totals_row + 1, column=2).value, self.result.totals.reservation_count)

    def test_xlsx_money_and_date_cells_are_typed_for_formulas(self):
        workbook = load_workbook(io.BytesIO(render_property_financial_report_xlsx(self.result)))
        sheet = workbook.active
        data_row = HEADER_ROW + 1

        gross_cell = sheet.cell(row=data_row, column=7)
        self.assertIsInstance(gross_cell.value, numbers.Number)
        self.assertNotIsInstance(gross_cell.value, str)
        self.assertEqual(gross_cell.value, 150)
        self.assertEqual(gross_cell.number_format, EXCEL_MONEY_NUMBER_FORMAT)

        check_in_cell = sheet.cell(row=data_row, column=3)
        self.assertIsInstance(check_in_cell.value, (date, datetime))
        self.assertEqual(check_in_cell.value.date() if isinstance(check_in_cell.value, datetime) else check_in_cell.value, date(2026, 3, 10))

        missing_commission_net = sheet.cell(row=HEADER_ROW + 2, column=9)
        self.assertIsNone(missing_commission_net.value)

    def test_xlsx_totals_use_numeric_money_cells(self):
        workbook = load_workbook(io.BytesIO(render_property_financial_report_xlsx(self.result)))
        sheet = workbook.active
        totals_row = HEADER_ROW + len(self.result.rows) + 2

        for column in (7, 8, 9):
            cell = sheet.cell(row=totals_row, column=column)
            self.assertIsInstance(cell.value, numbers.Number)
            self.assertNotIsInstance(cell.value, str)
            self.assertEqual(cell.number_format, EXCEL_MONEY_NUMBER_FORMAT)
